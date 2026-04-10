import argparse
import sqlite3
from pathlib import Path
from typing import Any
from urllib.parse import urlsplit

from openpyxl import load_workbook

from backend.config import settings
from backend.database import init_db, utcnow_iso
from backend.domain_types import LeadStatus, normalize_email, normalize_website


OUTREACH_TABLES = [
    "email_events",
    "email_variants",
    "replies",
    "seo_insights",
    "experiment_results",
    "experiments",
    "leads",
]

REQUIRED_COLUMNS = ["company", "contact_email", "website", "segment"]


def _clean(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _normalize_status(value: str | None) -> str:
    if not value:
        return LeadStatus.NEW.value
    lowered = value.strip().lower()
    if lowered in {LeadStatus.NEW.value, LeadStatus.WRITTEN.value, LeadStatus.CONTACTED.value}:
        return lowered
    raise ValueError(f"Unsupported lead status '{value}'. Allowed: new, written, contacted")


def _is_http_url(value: str) -> bool:
    parts = urlsplit(value)
    return parts.scheme.lower() in {"http", "https"} and bool(parts.netloc)


def reset_tables(conn: sqlite3.Connection) -> None:
    for table in OUTREACH_TABLES:
        conn.execute(f"DELETE FROM {table}")

    placeholders = ",".join("?" for _ in OUTREACH_TABLES)
    conn.execute(f"DELETE FROM sqlite_sequence WHERE name IN ({placeholders})", OUTREACH_TABLES)


def import_leads(conn: sqlite3.Connection, xlsx_path: Path) -> tuple[int, list[str]]:
    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        header = [str(h).strip() if h is not None else "" for h in header_row]
        idx = {name: i for i, name in enumerate(header)}

        missing = [col for col in REQUIRED_COLUMNS if col not in idx]
        if missing:
            raise ValueError(f"Missing required Excel columns: {missing}")

        imported = 0
        seen_segments: set[str] = set()

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not any(v not in (None, "") for v in row):
                continue

            company = _clean(row[idx["company"]])
            contact_email = _clean(row[idx["contact_email"]])
            website = _clean(row[idx["website"]])
            segment = _clean(row[idx["segment"]])
            industry = _clean(row[idx.get("industry", -1)]) if "industry" in idx else None
            country = _clean(row[idx.get("country", -1)]) if "country" in idx else None
            source = _clean(row[idx.get("source", -1)]) if "source" in idx else None
            status = _clean(row[idx.get("status", -1)]) if "status" in idx else LeadStatus.NEW.value

            if not company or not contact_email or not website or not segment:
                raise ValueError(f"Row has missing required values: {row}")
            if not _is_http_url(website):
                raise ValueError(f"Website must include http/https URL format: {website}")

            normalized_email = normalize_email(contact_email)
            normalized_website = normalize_website(website)
            now = utcnow_iso()
            try:
                conn.execute(
                    """
                    INSERT INTO leads (company, contact_email, website, segment, industry, country, source, created_at, status)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        company,
                        normalized_email,
                        normalized_website,
                        segment,
                        industry,
                        country,
                        source,
                        now,
                        _normalize_status(status),
                    ),
                )
            except sqlite3.IntegrityError:
                # Skip duplicate leads based on normalized contact_email + website dedupe rule.
                continue
            seen_segments.add(segment)
            imported += 1

        return imported, sorted(seen_segments)
    finally:
        workbook.close()


def seed_experiments(conn: sqlite3.Connection, segments: list[str]) -> int:
    angle_by_segment = {
        "restaurant": "reservation intent visibility",
        "bakery_desserts": "local foot traffic",
        "cafe_bar": "near-me demand capture",
        "coffee_tea": "menu page discoverability",
        "specialty_food_retail": "high-intent product pages",
    }

    now = utcnow_iso()
    count = 0
    for segment in segments:
        angle = angle_by_segment.get(segment, "local visibility growth")
        subject = f"Case-based growth idea for {segment}"
        conn.execute(
            """
            INSERT INTO experiments (segment, messaging_angle, email_format, subject_variant, created_at, active)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (segment, angle, "short", subject, now, 1),
        )
        count += 1

    return count


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Reset outreach tables, import leads from Excel, and seed baseline experiments."
    )
    parser.add_argument("--db-path", default=str(settings.db_path), help="Path to SQLite database")
    parser.add_argument(
        "--xlsx-path",
        default=str(settings.leads_xlsx_path),
        help="Path to leads Excel file",
    )
    return parser


def main() -> None:
    args = build_parser().parse_args()
    db_path = Path(args.db_path)
    xlsx_path = Path(args.xlsx_path)

    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON;")

    try:
        init_db(conn)
        conn.execute("BEGIN")
        reset_tables(conn)
        imported, segments = import_leads(conn, xlsx_path)
        seeded = seed_experiments(conn, segments)
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()

    print(f"Reset completed on: {db_path}")
    print(f"Imported leads: {imported}")
    print(f"Segments discovered: {segments}")
    print(f"Seeded experiments: {seeded}")


if __name__ == "__main__":
    main()
